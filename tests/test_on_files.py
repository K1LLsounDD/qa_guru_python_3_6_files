import csv
import os
import zipfile
from zip_func import create_arch, delete_arch
import PyPDF2
from openpyxl import load_workbook

path_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.path.pardir, 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.path.pardir, 'resources')
zip_archive = os.path.join(path_resources, 'archive.zip')


def test_read_pdf():
    create_arch(path_files, zip_archive)

    with zipfile.ZipFile(zip_archive) as zp:
        pdf_file = zp.extract('rekomend.pdf')
        reader = PyPDF2.PdfReader(pdf_file)
        pages = reader.pages[0]
        text = pages.extract_text()
        assert 'Выдержка  из  " Методических  рекомендаций ..."' in text
    os.remove('rekomend.pdf')


def test_read_csv():
    delete_arch(zip_archive)
    create_arch(path_files, zip_archive)

    with zipfile.ZipFile(zip_archive) as cvs:
        csv_file = cvs.extract('orders.csv')
        with open(csv_file) as csvfile:
            csvfile = csv.reader(csvfile)
            csv_list = []
            for r in csvfile:
                text = ' '.join(r).replace(';', ' ')
                csv_list.append(text)
        assert csv_list[1] == '35622 Иванов Отклонен'
    os.remove('orders.csv')


def test_read_xlsx():
    delete_arch(zip_archive)
    create_arch(path_files, zip_archive)

    with zipfile.ZipFile(zip_archive) as xlsx_:
        xlsx_file = xlsx_.extract('file_example_XLSX_50.xlsx')
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active
        assert sheet.cell(row=4, column=2).value == 'Philip'
    os.remove('file_example_XLSX_50.xlsx')
    delete_arch(zip_archive)
